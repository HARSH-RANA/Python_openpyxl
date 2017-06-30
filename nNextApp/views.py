from django.shortcuts import render
from django.http import HttpResponse, HttpResponseBadRequest
from django import forms
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook



class UploadFileForm(forms.Form):
    file = forms.FileField()

def index(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            filehandle = request.FILES['file']
            return excel.make_response(filehandle.get_sheet(), "csv",
                                       file_name="download")
    else:
        form = UploadFileForm()
    return render(
        request,
        'nNextApp/upload_form.html',
        {
            'form': form,
            'title': 'Excel file upload and download',
            'header': ('Upload excel file ' +
                       'from your repository:')
        })

def import_data(request):

    if request.method=="POST":
      form = UploadFileForm(request.POST,
                              request.FILES)
      if form.is_valid():

       
        
        wb = load_workbook(request.FILES['file'],data_only=True)
        nows=0          # number of worksheets
        for i in wb.worksheets:
          nows=nows+1
        if(nows==2):

            ws = wb.worksheets[0]
            ws1 = wb.worksheets[1]
            # ws2 = wb.create_sheet("My", 0)
            
            w,h=ws.max_row,ws.max_column
            w1,h1=ws1.max_row,ws1.max_column
            
            dialer_arr=[None]*ws.max_column
            for i in range(ws.max_column):
              dialer_arr[i]=ws.cell(chr(65+i)+str(1)).value


        elif(nows==3):       

          ws0=wb.worksheets[0]
          ws = wb.worksheets[1]
          ws1 = wb.worksheets[2]
          
          
          w,h=ws.max_row,ws.max_column
          w1,h1=ws1.max_row,ws1.max_column
          
          #adding dispositions and bifurcation in array
          desp_arr=[None]*ws0.max_column
          for i in range(ws0.max_column):
            desp_arr[i]=ws0.cell(chr(65+i)+str(1)).value

          desp={}
          # insert predefined desposition in desp dictionary
          

          for i in range(2,ws0.max_row):
            desp[ws0.cell(chr(65+desp_arr.index('Dispositions'))+str(i)).value]=ws0.cell(chr(65+desp_arr.index('Bifurcation'))+str(i)).value

          dialer_arr=[None]*ws.max_column
          count=0
          status=""
          for i in range(ws.max_column):
            dialer_arr[i]=ws.cell(chr(65+i)+str(1)).value
          for i in range(ws.max_row):
            for j in range(6):
              count=count+1
              try:
                #status=desp[ws.cell(chr(65+dialer_arr.index('Disposition1')+j)+str(i+2)).value]

              
                if(desp[ws.cell(chr(65+dialer_arr.index('Disposition1')+j)+str(i+2)).value]=="Paid"):
                  ws[chr(65+dialer_arr.index('Disposition1')+6+j)+str(i+2)]="Paid"
                  ws[chr(65+dialer_arr.index('Check'))+str(i+2)]="Paid"
                  break
                else:
                    ws[chr(65+dialer_arr.index('Disposition1')+6+j)+str(i+2)]="Not Paid"
                    if(count==6):
                      ws[chr(65+dialer_arr.index('Check'))+str(i+2)]="Not Paid"
              except KeyError:
                ws[chr(65+dialer_arr.index('Disposition1')+6+j)+str(i+2)]="NA"
                # if(count==6):
                #   ws[chr(65+dialer_arr.index('Check'))+str(i+2)]="NA"

        else:
          html='<html><body><p>Please upload file having 2 or 3 worksheets.</p></body></html>'
          return HttpResponse(html)
        phone_Status={}
        for row in range(1,ws.max_row):
          status=str(ws.cell(chr(dialer_arr.index('Check')+65)+str(row+1)).value)
          if(status=="Paid"):
            phone_Status[str(ws.cell(chr(65+dialer_arr.index('phone_number'))+str(row+1)).value)]="Paid"    #add phones in dictionary
          #ws[chr(h+65)+str(row+1)]=str(ws.cell(chr(1+65)+str(row+1)).value)+status
        
        nnext_arr=[None]*ws1.max_column
        for i in range(ws1.max_column):
          nnext_arr[i]=ws1.cell(chr(65+i)+str(1)).value
        for row in range(1,ws1.max_row):
          try:
            ws1[chr(65+nnext_arr.index('comment_date'))+str(row+1)]=phone_Status[str(ws1.cell(chr(65+nnext_arr.index("mobile"))+str(row+1)).value)]
          except KeyError:
              ws1[chr(65+nnext_arr.index('comment_date'))+str(row+1)]="Not Paid"
          

        ws2=wb.create_sheet('pivot',0) 
        ws2[chr(65+0)+str(4)]="source"
        ws2[chr(65+1)+str(4)]="Not Paid"
        ws2[chr(65+2)+str(4)]="Paid"
        ws2[chr(65+3)+str(4)]="Total"

        pivot={}
        
        #concate publisher and status for key and count of status as value then inserting in dictionary
        for row in range(1,ws1.max_row):
          try:
            pivot[(str(ws1.cell(chr(65+nnext_arr.index('source'))+str(row+1)).value)+"_"+str(ws1.cell(chr(65+nnext_arr.index('comment_date'))+str(row+1)).value)).strip()]=pivot[(str(ws1.cell(chr(65+nnext_arr.index('source'))+str(row+1)).value)+"_"+str(ws1.cell(chr(65+nnext_arr.index('comment_date'))+str(row+1)).value)).strip()]+1
          except KeyError:
            pivot[(str(ws1.cell(chr(65+nnext_arr.index('source'))+str(row+1)).value)+"_"+str(ws1.cell(chr(65+nnext_arr.index('comment_date'))+str(row+1)).value)).strip()]=1
        
        keys=list(pivot.keys())    # create list of publisher_status 

        
        check=list()
        n=0
        for key in keys:
          if(key.endswith('_Paid')):
            if(key[:-5] in check):
              continue
            try:
              n=n+1
              ws2[chr(65)+str(n+4)]=key[:-5]      #keys.index(key)
              ws2[chr(65+2)+str(n+4)]=pivot[key]
              ws2[chr(65+1)+str(n+4)]=pivot[key[:-5]+"_Not Paid"]
              ws2[chr(65+3)+str(n+4)]=pivot[key]+pivot[key[:-5]+"_Not Paid"]
              # del pivot[key[:-5]+"_Not Paid"]
              # del keys[key[:-5]+"_Not Paid"]
              check.append(key[:-5])
            except KeyError:
              ws2[chr(65+3)+str(n+4)]=pivot[key]
          if(key.endswith('_Not Paid')):
            if(key[:-9] in check):
              continue
            try:
              n=n+1
              ws2[chr(65)+str(n+4)]=key[:-9]
              ws2[chr(65+1)+str(n+4)]=pivot[key]
              ws2[chr(65+2)+str(n+4)]=pivot[key[:-9]+"_Paid"]
              ws2[chr(65+3)+str(n+4)]=pivot[key]+pivot[key[:-9]+"_Paid"]
              # del pivot[key[:-5]+"_Paid"]
              # del keys[key[:-5]+"_Paid"]
              check.append(key[:-9])
            except KeyError:
              ws2[chr(65+3)+str(n+4)]=pivot[key]

        ws2[chr(65+0)+str(n+5)]="Grand Total"
        ws2[chr(65+1)+str(n+5)]=0
        ws2[chr(65+2)+str(n+5)]=0
        ws2[chr(65+3)+str(n+5)]=0
        for i in range(n):
          if(ws2.cell(chr(65+1)+str(5+i)).value==None):
            ws2[chr(65+1)+str(n+5)]=ws2.cell(chr(65+1)+str(n+5)).value+0
          else:
            ws2[chr(65+1)+str(n+5)]=ws2.cell(chr(65+1)+str(n+5)).value+ws2.cell(chr(65+1)+str(5+i)).value
          if(ws2.cell(chr(65+2)+str(5+i)).value==None):
            ws2[chr(65+2)+str(n+5)]=ws2.cell(chr(65+2)+str(n+5)).value+0  
          else:
            ws2[chr(65+2)+str(n+5)]=ws2.cell(chr(65+2)+str(n+5)).value+ws2.cell(chr(65+2)+str(5+i)).value
          if(ws2.cell(chr(65+3)+str(5+i)).value==None):
            ws2[chr(65+3)+str(n+5)]=ws2.cell(chr(65+3)+str(n+5)).value+0
          else:  
            ws2[chr(65+3)+str(n+5)]=ws2.cell(chr(65+3)+str(n+5)).value+ws2.cell(chr(65+3)+str(5+i)).value        

        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=pivot.xlsx'
        return response
      else:
        return HttpResponseBadRequest()
    else:
        form = UploadFileForm()
    return render(
        request,
        'upload_form.html',
        {
            'form': form,
            'title': 'Import excel data into database',
            'header': 'Please upload sample-data.xls:'
        })
            